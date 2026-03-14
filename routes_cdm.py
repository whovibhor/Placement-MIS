# routes_cdm.py — Company Data Management routes
import re
from datetime import datetime, date
from decimal import Decimal
from io import BytesIO

from flask import (
    flash, jsonify, redirect, render_template, request, send_file, url_for,
)
from mysql.connector import Error

from helpers import (
    get_connection, normalize_rows, invalidate_analytics_cache,
)


# ── CDM constants ────────────────────────────────────────────────────────────
CDM_EXCEL_HEADERS = {
    "Company ID": "company_id",
    "Company Name": "company_name",
    "Date JD Received": "jd_received_date",
    "Data Shared(Y/N)": "data_shared",
    "Process Date": "process_date",
    "Recieved By": "received_by",
    "Course": "course",
    "Position Offered": "role",
    "CTC": "ctc_text",
    "HR POC Name": "hr_name",
    "HR POC Designation": "hr_designation",
    "HR POC Email": "hr_email",
    "HR POC Phone": "hr_phone",
    "Process Mode(On-Campus / Virtual)": "process_mode",
    "Location": "location",
    "Notes": "notes",
}

CDM_EDITABLE_FIELDS = {
    "role", "ctc_text", "jd_received_date", "process_date", "data_shared",
    "location", "notes", "status", "courses",
    "jd_briefing_done", "jd_briefing_date", "jd_briefing_conducted_by",
}

VALID_DRIVE_TYPES = {"Mandatory", "Interest Based", "Core"}


def parse_courses_value(value):
    if value is None:
        return []
    if isinstance(value, list):
        raw_items = value
    else:
        raw_items = str(value).split(",")
    cleaned = []
    seen = set()
    for item in raw_items:
        course = str(item).strip()
        if course and course.lower() != "nan" and course.lower() not in seen:
            cleaned.append(course)
            seen.add(course.lower())
    return cleaned


def parse_drive_courses(value):
    entries = []
    seen = set()
    if value is None:
        return entries
    raw_items = value if isinstance(value, list) else str(value).split(",")
    for item in raw_items:
        if isinstance(item, dict):
            course = str(item.get("course_name") or "").strip()
            drive_type = (item.get("drive_type") or "").strip() or None
        else:
            course = str(item).strip()
            drive_type = None
        if not course or course.lower() == "nan":
            continue
        if drive_type and drive_type not in VALID_DRIVE_TYPES:
            drive_type = None
        key = course.lower()
        if key in seen:
            continue
        seen.add(key)
        entries.append({"course_name": course, "drive_type": drive_type})
    return entries


def format_course_label(course_name, drive_type):
    if drive_type:
        return f"{course_name} ({drive_type})"
    return course_name


# ── CDM helpers ──────────────────────────────────────────────────────────────
def generate_company_id(name, cursor):
    clean = ''.join(c for c in name.upper() if c.isalnum())
    if len(clean) >= 4:
        cid = clean[:2] + clean[-2:]
    else:
        cid = (clean + "XXXX")[:4]
    base = cid
    i = 1
    while True:
        cursor.execute("SELECT 1 FROM companies WHERE company_id=%s", (cid,))
        if not cursor.fetchone():
            return cid
        cid = f"{base}{i}"
        i += 1


def parse_date_field(val):
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def parse_ctc_value(ctc_text):
    if not ctc_text:
        return None
    text = str(ctc_text).strip()
    parts = re.split(r'\s*[-&,/]\s*', text)
    for part in parts:
        m = re.search(r'(\d+\.?\d*)', part)
        if m:
            val = float(m.group(1))
            if val > 0:
                return val
    return None


def get_request_actor():
    actor = (
        request.headers.get("X-Actor")
        or request.headers.get("X-User")
        or request.args.get("actor")
    )
    if not actor and request.is_json:
        payload = request.get_json(silent=True) or {}
        actor = payload.get("actor")
    actor = (actor or "system").strip()
    return actor[:100] if actor else "system"


def sync_student_placed_status(cursor, reg_no, company_name, role, ctc_text, to_status):
    status = (to_status or "").strip().lower()
    if status == "placed":
        parsed_ctc = parse_ctc_value(ctc_text)
        cursor.execute(
            "UPDATE students "
            "SET status='Placed', "
            "company_name = COALESCE(NULLIF(company_name, ''), %s), "
            "designation = COALESCE(NULLIF(designation, ''), %s), "
            "ctc = COALESCE(ctc, %s), "
            "placed_date = COALESCE(placed_date, CURDATE()) "
            "WHERE reg_no = %s",
            (company_name, role, parsed_ctc, reg_no),
        )


def append_round_transition_log(
    cursor,
    drive_id,
    reg_no,
    from_round,
    to_round,
    from_status,
    to_status,
    actor,
    note=None,
):
    cursor.execute(
        "INSERT INTO drive_round_transitions "
        "(drive_id, reg_no, from_round, to_round, from_status, to_status, actor, note, changed_at) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())",
        (
            drive_id,
            reg_no,
            from_round,
            to_round,
            from_status,
            to_status,
            actor,
            (note or None),
        ),
    )


def register_cdm_routes(app):
    """Register all CDM routes on the Flask app."""

    # ── CDM main page ────────────────────────────────────────────────────
    @app.route("/cdm")
    def cdm_page():
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT d.drive_id, d.company_id, c.company_name, d.role, d.ctc_text,
                       d.jd_received_date, d.process_date, d.data_shared,
                       d.location, d.notes, d.status,
                      d.jd_briefing_done, d.jd_briefing_date, d.jd_briefing_conducted_by,
                      c.received_by
                FROM company_drives d
                JOIN companies c ON d.company_id = c.company_id
                ORDER BY d.drive_id DESC
            """)
            drives = cursor.fetchall()
            normalize_rows(drives)

            drive_ids = [d["drive_id"] for d in drives]
            course_map = {}
            if drive_ids:
                format_ids = ",".join(["%s"] * len(drive_ids))
                cursor.execute(
                    f"SELECT drive_id, course_name, drive_type FROM drive_courses WHERE drive_id IN ({format_ids})",
                    tuple(drive_ids),
                )
                for row in cursor.fetchall():
                    course_map.setdefault(row["drive_id"], []).append(
                        format_course_label(row["course_name"], row.get("drive_type"))
                    )
            for d in drives:
                d["courses"] = ", ".join(course_map.get(d["drive_id"], []))
                d["data_shared"] = "Yes" if d.get("data_shared") else "No"
                d["jd_briefing_done"] = "Yes" if d.get("jd_briefing_done") else "No"

            cursor.execute("""
                SELECT c.company_id, c.company_name, c.received_by,
                       COUNT(DISTINCT d.drive_id) AS drive_count,
                       MAX(d.process_date) AS latest_process_date,
                       COUNT(DISTINCT ds.reg_no) AS participants_count
                FROM companies c
                LEFT JOIN company_drives d ON c.company_id = d.company_id
                LEFT JOIN drive_students ds ON d.drive_id = ds.drive_id
                GROUP BY c.company_id, c.company_name, c.received_by
                ORDER BY latest_process_date DESC, c.company_id DESC
            """)
            companies = cursor.fetchall()
            normalize_rows(companies)

            cursor.close()
            conn.close()
        except Error as e:
            drives = []
            companies = []
            flash(f"Database error: {e}", "danger")
        return render_template("cdm.html", drives_json=drives, companies_json=companies)

    # ── CDM JSON endpoint ────────────────────────────────────────────────
    @app.route("/api/cdm")
    def api_cdm():
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT d.drive_id, d.company_id, c.company_name, d.role, d.ctc_text,
                       d.jd_received_date, d.process_date, d.data_shared,
                       d.location, d.notes, d.status,
                      d.jd_briefing_done, d.jd_briefing_date, d.jd_briefing_conducted_by,
                      c.received_by
                FROM company_drives d
                JOIN companies c ON d.company_id = c.company_id
                ORDER BY d.drive_id DESC
            """)
            drives = cursor.fetchall()
            normalize_rows(drives)
            drive_ids = [d["drive_id"] for d in drives]
            course_map = {}
            if drive_ids:
                format_ids = ",".join(["%s"] * len(drive_ids))
                cursor.execute(
                    f"SELECT drive_id, course_name, drive_type FROM drive_courses WHERE drive_id IN ({format_ids})",
                    tuple(drive_ids),
                )
                for row in cursor.fetchall():
                    course_map.setdefault(row["drive_id"], []).append(
                        format_course_label(row["course_name"], row.get("drive_type"))
                    )
            for d in drives:
                d["courses"] = ", ".join(course_map.get(d["drive_id"], []))
                d["data_shared"] = "Yes" if d.get("data_shared") else "No"
                d["jd_briefing_done"] = "Yes" if d.get("jd_briefing_done") else "No"
            cursor.close()
            conn.close()
            return jsonify({"data": drives})
        except Error as e:
            return jsonify({"data": [], "error": str(e)}), 500

    # ── CDM Import ───────────────────────────────────────────────────────
    @app.route("/cdm/import", methods=["GET", "POST"])
    def cdm_import():
        if request.method == "GET":
            return redirect(url_for("upload"))

        import pandas as pd

        file = request.files.get("file")
        if not file or file.filename == "":
            flash("No file selected.", "danger")
            return redirect(url_for("cdm_import"))
        if not file.filename.lower().endswith(".xlsx"):
            flash("Only .xlsx files are accepted.", "danger")
            return redirect(url_for("cdm_import"))

        try:
            df = pd.read_excel(file, engine="openpyxl")
        except Exception as e:
            flash(f"Could not read Excel file: {e}", "danger")
            return redirect(url_for("cdm_import"))

        df.columns = [str(h).strip() for h in df.columns]
        df = df.where(pd.notnull(df), None)

        col_map = {}
        for excel_h, internal in CDM_EXCEL_HEADERS.items():
            for col in df.columns:
                if col.strip().lower() == excel_h.strip().lower():
                    col_map[col] = internal
                    break
        df.rename(columns=col_map, inplace=True)

        required = {"company_name"}
        present = set(df.columns)
        missing = required - present
        if missing:
            flash(f"Missing required columns: {missing}", "danger")
            return redirect(url_for("cdm_import"))

        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)

            companies_added = 0
            drives_added = 0
            hr_added = 0

            for _, row in df.iterrows():
                company_name = str(row.get("company_name", "")).strip() if row.get("company_name") else None
                if not company_name or company_name.lower() == "nan":
                    continue

                company_id = str(row.get("company_id", "")).strip() if row.get("company_id") else None
                if not company_id or company_id.lower() == "nan":
                    company_id = generate_company_id(company_name, cursor)

                cursor.execute("SELECT 1 FROM companies WHERE company_id=%s", (company_id,))
                if not cursor.fetchone():
                    cursor.execute(
                        "INSERT INTO companies (company_id, company_name, received_by) VALUES (%s, %s, %s)",
                        (company_id, company_name, None),
                    )
                    companies_added += 1

                role = str(row.get("role", "")).strip() if row.get("role") else None
                if role and role.lower() == "nan":
                    role = None
                ctc_text = str(row.get("ctc_text", "")).strip() if row.get("ctc_text") else None
                if ctc_text and ctc_text.lower() == "nan":
                    ctc_text = None
                jd_date = parse_date_field(row.get("jd_received_date"))
                proc_date = parse_date_field(row.get("process_date"))
                data_shared_raw = str(row.get("data_shared", "")).strip().upper() if row.get("data_shared") else ""
                data_shared = True if data_shared_raw in ("Y", "YES", "TRUE", "1") else False
                location = str(row.get("location", "")).strip() if row.get("location") else None
                if location and location.lower() == "nan":
                    location = None
                received_by = str(row.get("received_by", "")).strip() if row.get("received_by") else None
                if received_by and received_by.lower() == "nan":
                    received_by = None
                if received_by:
                    cursor.execute(
                        "UPDATE companies SET received_by = COALESCE(NULLIF(received_by, ''), %s) WHERE company_id=%s",
                        (received_by, company_id),
                    )
                notes = str(row.get("notes", "")).strip() if row.get("notes") else None
                if notes and notes.lower() == "nan":
                    notes = None

                dup_check_sql = (
                    "SELECT drive_id FROM company_drives "
                    "WHERE company_id=%s AND role=%s"
                )
                dup_params = [company_id, role]
                if proc_date:
                    dup_check_sql += " AND process_date=%s"
                    dup_params.append(proc_date)
                else:
                    dup_check_sql += " AND process_date IS NULL"
                cursor.execute(dup_check_sql, tuple(dup_params))
                if cursor.fetchone():
                    continue

                cursor.execute(
                    "INSERT INTO company_drives "
                    "(company_id, role, ctc_text, jd_received_date, process_date, "
                    "data_shared, location, notes, status, jd_briefing_done, jd_briefing_date, jd_briefing_conducted_by) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (company_id, role, ctc_text, jd_date, proc_date,
                     data_shared, location, notes, "Upcoming", False, None, None),
                )
                drive_id = cursor.lastrowid
                drives_added += 1

                courses_raw = str(row.get("course", "")).strip() if row.get("course") else ""
                if courses_raw and courses_raw.lower() != "nan":
                    courses = [c.strip() for c in courses_raw.split(",") if c.strip()]
                    if courses:
                        cursor.executemany(
                            "INSERT IGNORE INTO drive_courses (drive_id, course_name) VALUES (%s, %s)",
                            [(drive_id, c) for c in courses],
                        )

                hr_name = str(row.get("hr_name", "")).strip() if row.get("hr_name") else None
                if hr_name and hr_name.lower() == "nan":
                    hr_name = None
                if hr_name:
                    hr_desig = str(row.get("hr_designation", "")).strip() if row.get("hr_designation") else None
                    if hr_desig and hr_desig.lower() == "nan":
                        hr_desig = None
                    hr_email = str(row.get("hr_email", "")).strip() if row.get("hr_email") else None
                    if hr_email and hr_email.lower() == "nan":
                        hr_email = None
                    hr_phone = str(row.get("hr_phone", "")).strip() if row.get("hr_phone") else None
                    if hr_phone and hr_phone.lower() == "nan":
                        hr_phone = None
                    cursor.execute(
                        "SELECT hr_id FROM company_hr WHERE company_id=%s AND name=%s",
                        (company_id, hr_name),
                    )
                    if not cursor.fetchone():
                        cursor.execute(
                            "INSERT INTO company_hr (company_id, name, designation, email, phone) "
                            "VALUES (%s, %s, %s, %s, %s)",
                            (company_id, hr_name, hr_desig, hr_email, hr_phone),
                        )
                        hr_added += 1

            conn.commit()
            cursor.close()
            conn.close()
            flash(
                f"CDM import successful — {companies_added} companies, "
                f"{drives_added} drives, {hr_added} HR contacts added.",
                "success",
            )
        except Error as e:
            flash(f"Database error: {e}", "danger")

        return redirect(url_for("upload"))

    # ── Company detail page ──────────────────────────────────────────────
    @app.route("/cdm/company/<company_id>")
    def cdm_company_detail(company_id):
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM companies WHERE company_id=%s", (company_id,))
            company = cursor.fetchone()
            if not company:
                flash("Company not found.", "danger")
                return redirect(url_for("cdm_page"))

            cursor.execute(
                "SELECT * FROM company_drives WHERE company_id=%s ORDER BY drive_id DESC",
                (company_id,),
            )
            drives = cursor.fetchall()
            normalize_rows(drives)

            for d in drives:
                cursor.execute(
                    "SELECT course_name, drive_type FROM drive_courses WHERE drive_id=%s",
                    (d["drive_id"],),
                )
                d["courses"] = ", ".join([
                    format_course_label(r["course_name"], r.get("drive_type")) for r in cursor.fetchall()
                ])
                d["data_shared"] = "Yes" if d.get("data_shared") else "No"
                cursor.execute(
                    "SELECT COUNT(*) AS cnt FROM drive_students WHERE drive_id=%s",
                    (d["drive_id"],),
                )
                d["student_count"] = cursor.fetchone()["cnt"]
                cursor.execute(
                    "SELECT COUNT(*) AS cnt FROM drive_rounds WHERE drive_id=%s",
                    (d["drive_id"],),
                )
                d["round_count"] = cursor.fetchone()["cnt"]

            cursor.execute(
                "SELECT * FROM company_hr WHERE company_id=%s ORDER BY name",
                (company_id,),
            )
            hr_contacts = cursor.fetchall()

            cursor.close()
            conn.close()
        except Error as e:
            flash(f"Database error: {e}", "danger")
            return redirect(url_for("cdm_page"))

        return render_template(
            "cdm_company.html",
            company=company,
            drives=drives,
            hr_contacts=hr_contacts,
        )

    # ── Drive CRUD ───────────────────────────────────────────────────────
    @app.route("/api/cdm/delete-drive/<int:drive_id>", methods=["POST"])
    def cdm_delete_drive(drive_id):
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM drive_students WHERE drive_id=%s", (drive_id,))
            cursor.execute("DELETE FROM drive_rounds WHERE drive_id=%s", (drive_id,))
            cursor.execute("DELETE FROM drive_courses WHERE drive_id=%s", (drive_id,))
            cursor.execute("DELETE FROM company_drives WHERE drive_id=%s", (drive_id,))
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/cdm/drive", methods=["POST"])
    def cdm_create_drive():
        data = request.get_json(silent=True) or {}
        company_id = (data.get("company_id") or "").strip()
        role = (data.get("role") or "").strip()
        if not company_id or not role:
            return jsonify({"ok": False, "error": "Company ID and role are required."}), 400
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT 1 FROM companies WHERE company_id=%s", (company_id,))
            if not cursor.fetchone():
                cursor.close()
                conn.close()
                return jsonify({"ok": False, "error": "Company not found."}), 404
            cursor.execute(
                "INSERT INTO company_drives (company_id, role, ctc_text, process_date, jd_received_date, "
                "data_shared, location, status, notes, jd_briefing_done, jd_briefing_date, jd_briefing_conducted_by) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (
                    company_id,
                    role,
                    data.get("ctc_text") or None,
                    parse_date_field(data.get("process_date")) if data.get("process_date") else None,
                    parse_date_field(data.get("jd_received_date")) if data.get("jd_received_date") else None,
                    True if str(data.get("data_shared") or "").strip().upper() in ("Y", "YES", "TRUE", "1") else False,
                    data.get("location") or None,
                    data.get("status") or "Upcoming",
                    data.get("notes") or None,
                    True if str(data.get("jd_briefing_done") or "").strip().upper() in ("Y", "YES", "TRUE", "1") else False,
                    parse_date_field(data.get("jd_briefing_date")) if data.get("jd_briefing_date") else None,
                    (data.get("jd_briefing_conducted_by") or "").strip() or None,
                ),
            )
            drive_id = cursor.lastrowid

            courses = parse_drive_courses(data.get("courses"))
            if courses:
                cursor.executemany(
                    "INSERT IGNORE INTO drive_courses (drive_id, course_name, drive_type) VALUES (%s, %s, %s)",
                    [(drive_id, c["course_name"], c.get("drive_type")) for c in courses],
                )

            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True, "drive_id": drive_id})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    # ── Drive inline editing ─────────────────────────────────────────────
    @app.route("/api/cdm/drive/<int:drive_id>", methods=["PUT"])
    def cdm_update_drive(drive_id):
        data = request.get_json()
        if not data or "field" not in data:
            return jsonify({"error": "Missing field"}), 400

        field = data["field"]
        value = data.get("value")

        if field not in CDM_EDITABLE_FIELDS:
            return jsonify({"error": f"Field '{field}' is not editable"}), 400

        if value is not None and str(value).strip() == "":
            value = None

        if field in ("jd_received_date", "process_date", "jd_briefing_date") and value:
            value = parse_date_field(value)

        if field in ("data_shared", "jd_briefing_done"):
            if isinstance(value, str):
                value = True if value.strip().upper() in ("Y", "YES", "TRUE", "1") else False

        if field == "status" and value:
            valid_statuses = {"Upcoming", "Ongoing", "Completed", "Cancelled"}
            if value not in valid_statuses:
                return jsonify({"error": f"Status must be one of: {', '.join(valid_statuses)}"}), 400

        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)

            if field == "courses":
                courses = parse_drive_courses(value)
                cursor.execute("SELECT 1 FROM company_drives WHERE drive_id=%s", (drive_id,))
                if not cursor.fetchone():
                    cursor.close()
                    conn.close()
                    return jsonify({"error": "Drive not found"}), 404
                cursor.execute("DELETE FROM drive_courses WHERE drive_id=%s", (drive_id,))
                if courses:
                    cursor.executemany(
                        "INSERT IGNORE INTO drive_courses (drive_id, course_name, drive_type) VALUES (%s, %s, %s)",
                        [(drive_id, c["course_name"], c.get("drive_type")) for c in courses],
                    )
                conn.commit()
                cursor.close()
                conn.close()
                return jsonify({
                    "ok": True,
                    "drive_id": drive_id,
                    "field": "courses",
                    "value": ", ".join([format_course_label(c["course_name"], c.get("drive_type")) for c in courses]),
                })

            cursor.execute("SELECT * FROM company_drives WHERE drive_id=%s", (drive_id,))
            row = cursor.fetchone()
            if not row:
                cursor.close()
                conn.close()
                return jsonify({"error": "Drive not found"}), 404

            old_value = row.get(field)
            if isinstance(old_value, Decimal):
                old_value = float(old_value)
            elif isinstance(old_value, (date, datetime)):
                old_value = old_value.strftime("%Y-%m-%d")

            cursor.execute(
                "SELECT c.company_name FROM company_drives d "
                "JOIN companies c ON d.company_id = c.company_id WHERE d.drive_id = %s",
                (drive_id,),
            )
            company_row = cursor.fetchone()
            company_name = company_row["company_name"] if company_row else ""

            cursor.execute(
                f"UPDATE company_drives SET `{field}` = %s WHERE drive_id = %s",
                (value, drive_id),
            )

            cursor.execute(
                "INSERT INTO cdm_edit_log (drive_id, company_name, field_name, old_value, new_value, changed_at) "
                "VALUES (%s, %s, %s, %s, %s, %s)",
                (drive_id, company_name, field,
                 str(old_value) if old_value is not None else None,
                 str(value) if value is not None else None,
                 datetime.now()),
            )

            conn.commit()
            cursor.close()
            conn.close()

            display_value = value
            if field in ("data_shared", "jd_briefing_done"):
                display_value = "Yes" if value else "No"

            return jsonify({"ok": True, "drive_id": drive_id, "field": field, "value": display_value})
        except Error as e:
            return jsonify({"error": str(e)}), 500

    # ── CDM Change Log API ───────────────────────────────────────────────
    @app.route("/api/cdm/edit-log")
    def api_cdm_edit_log():
        limit = request.args.get("limit", 200, type=int)
        page = request.args.get("page", 1, type=int)
        drive_id = request.args.get("drive_id", None, type=int)
        offset = (page - 1) * limit
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)

            if drive_id:
                cursor.execute(
                    "SELECT COUNT(*) as cnt FROM cdm_edit_log WHERE drive_id = %s", (drive_id,)
                )
            else:
                cursor.execute("SELECT COUNT(*) as cnt FROM cdm_edit_log")
            total_count = cursor.fetchone()["cnt"]

            if drive_id:
                cursor.execute(
                    "SELECT * FROM cdm_edit_log WHERE drive_id = %s ORDER BY changed_at DESC LIMIT %s OFFSET %s",
                    (drive_id, limit, offset),
                )
            else:
                cursor.execute(
                    "SELECT * FROM cdm_edit_log ORDER BY changed_at DESC LIMIT %s OFFSET %s",
                    (limit, offset),
                )
            rows = cursor.fetchall()
            for r in rows:
                if r.get("changed_at"):
                    r["changed_at"] = r["changed_at"].strftime("%d %b %Y, %I:%M %p")
            cursor.close()
            conn.close()

            total_pages = (total_count + limit - 1) // limit if limit > 0 else 1
            return jsonify({
                "data": rows,
                "page": page,
                "total_pages": total_pages,
                "total_count": total_count,
            })
        except Error as e:
            return jsonify({"data": [], "error": str(e)}), 500

    # ── HR Management API ────────────────────────────────────────────────
    @app.route("/api/cdm/hr/<int:hr_id>", methods=["PUT"])
    def cdm_update_hr(hr_id):
        data = request.get_json()
        if not data or "field" not in data:
            return jsonify({"error": "Missing field"}), 400

        field = data["field"]
        value = data.get("value")
        allowed = {"name", "designation", "email", "phone"}

        if field not in allowed:
            return jsonify({"error": f"Field '{field}' is not editable"}), 400

        if value is not None and str(value).strip() == "":
            value = None

        if field == "email" and value:
            if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', str(value)):
                return jsonify({"error": "Invalid email format"}), 400

        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM company_hr WHERE hr_id=%s", (hr_id,))
            if not cursor.fetchone():
                cursor.close()
                conn.close()
                return jsonify({"error": "HR contact not found"}), 404

            cursor.execute(
                f"UPDATE company_hr SET `{field}` = %s WHERE hr_id = %s",
                (value, hr_id),
            )
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True, "hr_id": hr_id, "field": field, "value": value})
        except Error as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/api/cdm/hr/<int:hr_id>", methods=["DELETE"])
    def cdm_delete_hr(hr_id):
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM company_hr WHERE hr_id=%s", (hr_id,))
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/cdm/hr", methods=["POST"])
    def cdm_add_hr():
        data = request.get_json()
        if not data or "company_id" not in data or "name" not in data:
            return jsonify({"error": "company_id and name are required"}), 400

        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO company_hr (company_id, name, designation, email, phone) "
                "VALUES (%s, %s, %s, %s, %s)",
                (data["company_id"], data["name"], data.get("designation"),
                 data.get("email"), data.get("phone")),
            )
            hr_id = cursor.lastrowid
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True, "hr_id": hr_id})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    # ── Company search ───────────────────────────────────────────────────
    @app.route("/api/cdm/search")
    def cdm_search():
        q = request.args.get("q", "").strip()
        if len(q) < 2:
            return jsonify({"results": []})
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            like = f"%{q}%"
            cursor.execute(
                "SELECT DISTINCT c.company_id, c.company_name, d.role, d.ctc_text, d.status "
                "FROM companies c "
                "LEFT JOIN company_drives d ON c.company_id = d.company_id "
                "WHERE c.company_name LIKE %s OR d.role LIKE %s OR d.location LIKE %s "
                "ORDER BY c.company_name LIMIT 15",
                (like, like, like),
            )
            rows = cursor.fetchall()
            cursor.close()
            conn.close()
            return jsonify({"results": rows})
        except Error as e:
            return jsonify({"results": [], "error": str(e)}), 500

    # ── Company CRUD ─────────────────────────────────────────────────────
    @app.route("/api/cdm/company", methods=["POST"])
    def cdm_create_company():
        data = request.get_json(silent=True) or {}
        name = (data.get("company_name") or "").strip()
        if not name:
            return jsonify({"ok": False, "error": "Company name is required."}), 400
        received_by = (data.get("received_by") or "").strip() or None
        notes = (data.get("notes") or "").strip() or None
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                "SELECT company_id FROM companies WHERE company_name = %s", (name,)
            )
            existing = cursor.fetchone()
            if existing:
                cursor.close()
                conn.close()
                return jsonify({"ok": False, "error": "Company already exists.", "company_id": existing["company_id"]}), 409
            cid = generate_company_id(name, cursor)
            cursor.execute(
                "INSERT INTO companies (company_id, company_name, received_by, notes) "
                "VALUES (%s, %s, %s, %s)",
                (cid, name, received_by, notes),
            )
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({
                "ok": True, "company_id": cid, "company_name": name,
                "received_by": received_by,
                "notes": notes,
            })
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/cdm/company/<company_id>", methods=["PUT"])
    def cdm_update_company(company_id):
        data = request.get_json(silent=True) or {}
        name = (data.get("company_name") or "").strip()
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                "SELECT company_name, received_by, notes "
                "FROM companies WHERE company_id=%s",
                (company_id,),
            )
            existing = cursor.fetchone()
            if not existing:
                cursor.close()
                conn.close()
                return jsonify({"ok": False, "error": "Company not found."}), 404
            if not name:
                name = existing["company_name"]
            if "received_by" in data:
                received_by = (data.get("received_by") or "").strip() or None
            else:
                received_by = existing.get("received_by")
            if "notes" in data:
                notes = (data.get("notes") or "").strip() or None
            else:
                notes = existing.get("notes")
            cursor.execute(
                "UPDATE companies SET company_name=%s, received_by=%s, notes=%s "
                "WHERE company_id=%s",
                (name, received_by, notes, company_id),
            )
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True, "company_name": name})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/cdm/company/<company_id>", methods=["DELETE"])
    def cdm_delete_company(company_id):
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(
                "DELETE ds FROM drive_students ds "
                "JOIN company_drives d ON ds.drive_id = d.drive_id "
                "WHERE d.company_id = %s",
                (company_id,),
            )
            cursor.execute(
                "DELETE dr FROM drive_rounds dr "
                "JOIN company_drives d ON dr.drive_id = d.drive_id "
                "WHERE d.company_id = %s",
                (company_id,),
            )
            cursor.execute(
                "DELETE dc FROM drive_courses dc "
                "JOIN company_drives d ON dc.drive_id = d.drive_id "
                "WHERE d.company_id = %s",
                (company_id,),
            )
            cursor.execute(
                "DELETE FROM company_drives WHERE company_id = %s", (company_id,)
            )
            cursor.execute(
                "DELETE FROM company_hr WHERE company_id = %s", (company_id,)
            )
            cursor.execute(
                "DELETE FROM companies WHERE company_id = %s", (company_id,)
            )
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    # ── CDM Analytics API ────────────────────────────────────────────────
    @app.route("/api/cdm/analytics")
    def cdm_analytics():
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)

            cursor.execute("SELECT COUNT(*) AS total_drives FROM company_drives")
            total_drives = cursor.fetchone()["total_drives"]
            cursor.execute("SELECT COUNT(*) AS total_companies FROM companies")
            total_companies = cursor.fetchone()["total_companies"]
            cursor.execute("SELECT COUNT(DISTINCT reg_no) AS total_participating_students FROM drive_students")
            total_participating_students = cursor.fetchone()["total_participating_students"]
            cursor.execute(
                "SELECT COUNT(DISTINCT reg_no) AS total_selected_students "
                "FROM drive_students WHERE status IN ('Selected', 'Placed')"
            )
            total_selected_students = cursor.fetchone()["total_selected_students"]

            cursor.execute(
                "SELECT IFNULL(status, 'Upcoming') AS status, COUNT(*) AS cnt "
                "FROM company_drives GROUP BY status"
            )
            status_dist = {r["status"]: r["cnt"] for r in cursor.fetchall()}

            cursor.execute("SELECT ctc_text FROM company_drives WHERE ctc_text IS NOT NULL AND ctc_text != ''")
            ctc_rows = cursor.fetchall()
            ctc_values = []
            for r in ctc_rows:
                val = parse_ctc_value(r["ctc_text"])
                if val is not None:
                    ctc_values.append(val)

            ctc_stats = {}
            if ctc_values:
                ctc_values.sort()
                ctc_stats["mean"] = round(sum(ctc_values) / len(ctc_values), 2)
                mid = len(ctc_values) // 2
                ctc_stats["median"] = round(
                    (ctc_values[mid - 1] + ctc_values[mid]) / 2 if len(ctc_values) % 2 == 0 else ctc_values[mid], 2
                )
                ctc_stats["highest"] = round(max(ctc_values), 2)
                ctc_stats["lowest"] = round(min(ctc_values), 2)
                ctc_stats["count"] = len(ctc_values)

            cursor.execute("""
                SELECT c.company_name, COUNT(ds.reg_no) AS selected_count
                FROM drive_students ds
                JOIN company_drives d ON ds.drive_id = d.drive_id
                JOIN companies c ON d.company_id = c.company_id
                WHERE ds.status IN ('Selected', 'Placed')
                GROUP BY c.company_id, c.company_name
                ORDER BY selected_count DESC
                LIMIT 10
            """)
            top_placed_companies = cursor.fetchall()

            cursor.execute("""
                SELECT c.company_name, d.ctc_text
                FROM company_drives d
                JOIN companies c ON d.company_id = c.company_id
                WHERE d.ctc_text IS NOT NULL AND d.ctc_text != ''
            """)
            ctc_company_rows = cursor.fetchall()
            company_ctc_map = {}
            for r in ctc_company_rows:
                val = parse_ctc_value(r["ctc_text"])
                if val is not None:
                    name = r["company_name"]
                    if name not in company_ctc_map or val > company_ctc_map[name]:
                        company_ctc_map[name] = val
            highest_ctc_companies = sorted(
                [{"company_name": k, "ctc": v} for k, v in company_ctc_map.items()],
                key=lambda x: x["ctc"], reverse=True
            )[:10]

            cursor.execute("""
                                SELECT c.received_by,
                       COUNT(DISTINCT d.drive_id) AS total_drives,
                       COUNT(DISTINCT d.company_id) AS companies_brought
                FROM company_drives d
                                JOIN companies c ON d.company_id = c.company_id
                                WHERE c.received_by IS NOT NULL AND c.received_by != ''
                                GROUP BY c.received_by
                ORDER BY companies_brought DESC
            """)
            team_base = cursor.fetchall()

            cursor.execute("""
                                SELECT c.received_by, COUNT(ds.reg_no) AS selections
                FROM company_drives d
                                JOIN companies c ON d.company_id = c.company_id
                JOIN drive_students ds ON d.drive_id = ds.drive_id
                WHERE ds.status IN ('Selected', 'Placed')
                                    AND c.received_by IS NOT NULL AND c.received_by != ''
                                GROUP BY c.received_by
            """)
            team_selections = {r["received_by"]: r["selections"] for r in cursor.fetchall()}

            cursor.execute("""
                                SELECT c.received_by, d.ctc_text
                FROM company_drives d
                                JOIN companies c ON d.company_id = c.company_id
                                WHERE c.received_by IS NOT NULL AND c.received_by != ''
                  AND d.ctc_text IS NOT NULL AND d.ctc_text != ''
            """)
            team_ctc_rows = cursor.fetchall()
            team_ctc_map = {}
            for r in team_ctc_rows:
                val = parse_ctc_value(r["ctc_text"])
                if val is not None:
                    team_ctc_map.setdefault(r["received_by"], []).append(val)

            team_performance = []
            for row in team_base:
                person = row["received_by"]
                ctc_list = team_ctc_map.get(person, [])
                team_performance.append({
                    "name": person,
                    "companies_brought": row["companies_brought"],
                    "total_drives": row["total_drives"],
                    "selections": team_selections.get(person, 0),
                    "avg_ctc": round(sum(ctc_list) / len(ctc_list), 2) if ctc_list else None,
                    "highest_ctc": round(max(ctc_list), 2) if ctc_list else None,
                })

            cursor.close()
            conn.close()

            return jsonify({
                "total_drives": total_drives,
                "total_companies": total_companies,
                "total_participating_students": total_participating_students,
                "total_selected_students": total_selected_students,
                "status_dist": status_dist,
                "ctc_stats": ctc_stats,
                "top_placed_companies": top_placed_companies,
                "highest_ctc_companies": highest_ctc_companies,
                "team_performance": team_performance,
            })
        except Error as e:
            return jsonify({"error": str(e)}), 500

    # ── Student Linking API ──────────────────────────────────────────────
    @app.route("/api/cdm/drive/<int:drive_id>/students", methods=["GET"])
    def cdm_drive_students(drive_id):
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                "SELECT ds.reg_no, ds.status, ds.current_round, s.student_name, s.course, s.department, s.email, s.mobile_number "
                "FROM drive_students ds "
                "JOIN students s ON ds.reg_no = s.reg_no "
                "WHERE ds.drive_id = %s ORDER BY s.student_name",
                (drive_id,),
            )
            rows = cursor.fetchall()
            cursor.close()
            conn.close()
            return jsonify({"students": rows})
        except Error as e:
            return jsonify({"students": [], "error": str(e)}), 500

    @app.route("/api/cdm/drive/<int:drive_id>/students", methods=["POST"])
    def cdm_link_student(drive_id):
        data = request.get_json()
        if not data or "reg_no" not in data:
            return jsonify({"error": "reg_no is required"}), 400
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(
                "INSERT IGNORE INTO drive_students (drive_id, reg_no, status, current_round) "
                "VALUES (%s, %s, %s, %s)",
                (drive_id, data["reg_no"], data.get("status", "Applied"), data.get("current_round", 0)),
            )
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/cdm/drive/<int:drive_id>/students/<reg_no>", methods=["PUT"])
    def cdm_update_drive_student(drive_id, reg_no):
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data"}), 400
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            actor = get_request_actor()

            cursor.execute(
                "SELECT ds.current_round, ds.status, d.role, d.ctc_text, c.company_name "
                "FROM drive_students ds "
                "JOIN company_drives d ON ds.drive_id = d.drive_id "
                "JOIN companies c ON d.company_id = c.company_id "
                "WHERE ds.drive_id=%s AND ds.reg_no=%s",
                (drive_id, reg_no),
            )
            before = cursor.fetchone()
            if not before:
                cursor.close()
                conn.close()
                return jsonify({"error": "Student not linked to drive"}), 404

            cursor.execute(
                "SELECT IFNULL(MAX(round_order), 0) AS max_round FROM drive_rounds WHERE drive_id=%s",
                (drive_id,),
            )
            max_round = int((cursor.fetchone() or {}).get("max_round") or 0)

            sets = []
            vals = []
            next_status = before.get("status")
            next_round = int(before.get("current_round") or 0)
            if "status" in data:
                incoming_status = str(data["status"] or "").strip()
                if incoming_status.lower() == "selected" and max_round and next_round >= max_round:
                    incoming_status = "Placed"
                sets.append("status = %s")
                vals.append(incoming_status)
                next_status = incoming_status
            if "current_round" in data:
                requested_round = int(data["current_round"] or 0)
                if max_round and requested_round > max_round:
                    requested_round = max_round
                sets.append("current_round = %s")
                vals.append(requested_round)
                next_round = requested_round
            if not sets:
                cursor.close()
                conn.close()
                return jsonify({"error": "Nothing to update"}), 400

            if (
                max_round
                and next_round >= max_round
                and str(next_status or "").strip().lower() == "selected"
            ):
                next_status = "Placed"
                if "status = %s" not in sets:
                    sets.append("status = %s")
                    vals.append(next_status)

            vals.extend([drive_id, reg_no])
            cursor.execute(
                f"UPDATE drive_students SET {', '.join(sets)} WHERE drive_id = %s AND reg_no = %s",
                tuple(vals),
            )

            from_round = int(before.get("current_round") or 0)
            from_status = before.get("status")
            if from_round != next_round or str(from_status or "") != str(next_status or ""):
                append_round_transition_log(
                    cursor,
                    drive_id,
                    reg_no,
                    from_round,
                    next_round,
                    from_status,
                    next_status,
                    actor,
                    note="single_update",
                )

            sync_student_placed_status(
                cursor,
                reg_no,
                before.get("company_name"),
                before.get("role"),
                before.get("ctc_text"),
                next_status,
            )

            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/cdm/drive/<int:drive_id>/students/rounds/bulk", methods=["PUT"])
    def cdm_bulk_update_student_rounds(drive_id):
        data = request.get_json(silent=True) or {}
        reg_nos = data.get("reg_nos") or []
        target_round = data.get("target_round")
        status = data.get("status")

        if not isinstance(reg_nos, list) or not reg_nos:
            return jsonify({"ok": False, "error": "reg_nos list is required"}), 400
        try:
            target_round = int(target_round)
        except Exception:
            return jsonify({"ok": False, "error": "target_round must be an integer"}), 400
        if target_round < 0:
            return jsonify({"ok": False, "error": "target_round must be >= 0"}), 400

        safe_reg_nos = [str(r).strip() for r in reg_nos if str(r).strip()]
        if not safe_reg_nos:
            return jsonify({"ok": False, "error": "No valid registration numbers provided"}), 400

        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            actor = get_request_actor()
            cursor.execute(
                "SELECT IFNULL(MAX(round_order), 0) AS max_round FROM drive_rounds WHERE drive_id=%s",
                (drive_id,),
            )
            max_round = int((cursor.fetchone() or {}).get("max_round") or 0)
            if target_round > max_round:
                target_round = max_round

            cursor.execute(
                "SELECT ds.reg_no, ds.current_round, ds.status, d.role, d.ctc_text, c.company_name "
                "FROM drive_students ds "
                "JOIN company_drives d ON ds.drive_id = d.drive_id "
                "JOIN companies c ON d.company_id = c.company_id "
                f"WHERE ds.drive_id = %s AND ds.reg_no IN ({','.join(['%s'] * len(safe_reg_nos))})",
                tuple([drive_id] + safe_reg_nos),
            )
            before_rows = cursor.fetchall()
            before_map = {r["reg_no"]: r for r in before_rows}

            normalized_status = status
            if normalized_status is not None:
                normalized_status = str(normalized_status).strip()
                if normalized_status.lower() == "selected" and max_round and target_round >= max_round:
                    normalized_status = "Placed"

            placeholders = ",".join(["%s"] * len(safe_reg_nos))
            vals = [target_round]
            set_parts = ["current_round = %s"]
            if normalized_status is not None:
                set_parts.append("status = %s")
                vals.append(normalized_status)
            vals.extend([drive_id] + safe_reg_nos)
            query = (
                f"UPDATE drive_students SET {', '.join(set_parts)} "
                f"WHERE drive_id = %s AND reg_no IN ({placeholders})"
            )
            cursor.execute(query, tuple(vals))
            updated = cursor.rowcount

            if updated:
                cursor.execute(
                    "SELECT reg_no, current_round, status FROM drive_students "
                    f"WHERE drive_id = %s AND reg_no IN ({placeholders})",
                    tuple([drive_id] + safe_reg_nos),
                )
                after_rows = cursor.fetchall()
                for after in after_rows:
                    reg_no = after["reg_no"]
                    before = before_map.get(reg_no)
                    if not before:
                        continue
                    from_round = int(before.get("current_round") or 0)
                    to_round = int(after.get("current_round") or 0)
                    from_status = before.get("status")
                    to_status = after.get("status")
                    if from_round != to_round or str(from_status or "") != str(to_status or ""):
                        append_round_transition_log(
                            cursor,
                            drive_id,
                            reg_no,
                            from_round,
                            to_round,
                            from_status,
                            to_status,
                            actor,
                            note="bulk_update",
                        )
                    sync_student_placed_status(
                        cursor,
                        reg_no,
                        before.get("company_name"),
                        before.get("role"),
                        before.get("ctc_text"),
                        to_status,
                    )

            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True, "updated": updated, "target_round": target_round})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/cdm/drive/<int:drive_id>/round-transitions")
    def cdm_drive_round_transitions(drive_id):
        reg_no = (request.args.get("reg_no") or "").strip()
        limit = request.args.get("limit", 100, type=int)
        if limit <= 0:
            limit = 100
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            if reg_no:
                cursor.execute(
                    "SELECT transition_id, drive_id, reg_no, from_round, to_round, from_status, to_status, actor, note, changed_at "
                    "FROM drive_round_transitions WHERE drive_id=%s AND reg_no=%s "
                    "ORDER BY changed_at DESC, transition_id DESC LIMIT %s",
                    (drive_id, reg_no, limit),
                )
            else:
                cursor.execute(
                    "SELECT transition_id, drive_id, reg_no, from_round, to_round, from_status, to_status, actor, note, changed_at "
                    "FROM drive_round_transitions WHERE drive_id=%s "
                    "ORDER BY changed_at DESC, transition_id DESC LIMIT %s",
                    (drive_id, limit),
                )
            rows = cursor.fetchall()
            normalize_rows(rows)
            cursor.close()
            conn.close()
            return jsonify({"transitions": rows})
        except Error as e:
            return jsonify({"transitions": [], "error": str(e)}), 500

    @app.route("/api/cdm/drive/<int:drive_id>/students/<reg_no>", methods=["DELETE"])
    def cdm_unlink_student(drive_id, reg_no):
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(
                "DELETE FROM drive_students WHERE drive_id = %s AND reg_no = %s",
                (drive_id, reg_no),
            )
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    # ── Bulk Student Import ──────────────────────────────────────────────
    @app.route("/api/cdm/drive/<int:drive_id>/students/bulk", methods=["POST"])
    def cdm_bulk_link_students(drive_id):
        if "file" not in request.files:
            return jsonify({"ok": False, "error": "No file uploaded"}), 400
        f = request.files["file"]
        if not f.filename or not f.filename.lower().endswith((".xlsx", ".xls")):
            return jsonify({"ok": False, "error": "Only .xlsx/.xls files accepted"}), 400
        try:
            import openpyxl

            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                return jsonify({"ok": False, "error": "Empty file"}), 400
            header_lower = [str(h).strip().lower() if h else "" for h in header]
            reg_col = None
            for i, h in enumerate(header_lower):
                if h in ("reg number", "reg_number", "reg no", "reg_no", "registration number", "registration_number"):
                    reg_col = i
                    break
            if reg_col is None:
                return jsonify({"ok": False, "error": "Column 'Reg Number' not found in header"}), 400
            file_regs = []
            for row in rows_iter:
                if row[reg_col] is not None:
                    val = str(row[reg_col]).strip()
                    if val:
                        file_regs.append(val)
            wb.close()
            if not file_regs:
                return jsonify({"ok": False, "error": "No registration numbers found in file"}), 400

            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT reg_no, student_name, course, department FROM students")
            all_students = cursor.fetchall()
            lookup = {}
            for s in all_students:
                lookup[s["reg_no"].upper()] = s

            linked = []
            not_found = []
            already = []
            for reg in file_regs:
                student = lookup.get(reg.upper())
                if not student:
                    not_found.append(reg)
                    continue
                try:
                    cursor.execute(
                        "INSERT INTO drive_students (drive_id, reg_no, status, current_round) "
                        "VALUES (%s, %s, %s, %s)",
                        (drive_id, student["reg_no"], "Applied", 0),
                    )
                    linked.append({
                        "reg_no": student["reg_no"],
                        "student_name": student["student_name"],
                        "course": student["course"],
                        "department": student["department"],
                    })
                except Exception:
                    already.append(student["reg_no"])
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({
                "ok": True,
                "linked": len(linked),
                "not_found": not_found,
                "already_linked": already,
                "students": linked,
            })
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/cdm/student-template")
    def cdm_student_template():
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(["Reg Number", "Student Name"])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="student_import_template.xlsx",
        )

    @app.route("/api/cdm/drive/<int:drive_id>/export.xlsx")
    def cdm_drive_export_excel(drive_id):
        try:
            import openpyxl

            conn = get_connection()
            cursor = conn.cursor(dictionary=True)

            cursor.execute(
                "SELECT d.*, c.company_name, c.received_by "
                "FROM company_drives d JOIN companies c ON d.company_id = c.company_id "
                "WHERE d.drive_id=%s",
                (drive_id,),
            )
            drive = cursor.fetchone()
            if not drive:
                cursor.close()
                conn.close()
                return jsonify({"error": "Drive not found"}), 404

            cursor.execute(
                "SELECT round_order, round_name, round_date "
                "FROM drive_rounds WHERE drive_id=%s ORDER BY round_order",
                (drive_id,),
            )
            rounds = cursor.fetchall()

            cursor.execute(
                "SELECT ds.reg_no, ds.status, ds.current_round, "
                "s.student_name, s.course, s.department, s.email, s.mobile_number "
                "FROM drive_students ds "
                "JOIN students s ON ds.reg_no = s.reg_no "
                "WHERE ds.drive_id=%s ORDER BY s.student_name",
                (drive_id,),
            )
            students = cursor.fetchall()

            cursor.close()
            conn.close()

            wb = openpyxl.Workbook()
            ws_info = wb.active
            ws_info.title = "Drive Overview"
            ws_info.append(["Field", "Value"])
            info_rows = [
                ("Drive ID", drive.get("drive_id")),
                ("Company", drive.get("company_name")),
                ("Role", drive.get("role")),
                ("CTC", drive.get("ctc_text")),
                ("Process Date", drive.get("process_date")),
                ("Venue", drive.get("location")),
                ("Status", drive.get("status")),
                ("JD Received", drive.get("jd_received_date")),
                ("Data Shared", "Yes" if drive.get("data_shared") else "No"),
                ("Received By", drive.get("received_by")),
            ]
            for key, value in info_rows:
                ws_info.append([key, value if value is not None else ""])

            ws_rounds = wb.create_sheet("Rounds")
            ws_rounds.append(["Round #", "Round Name", "Date", "Qualified Count"])
            for r in rounds:
                r_order = int(r.get("round_order") or 0)
                q_count = len([s for s in students if int(s.get("current_round") or 0) >= r_order])
                ws_rounds.append([r_order, r.get("round_name"), r.get("round_date"), q_count])

            ws_students = wb.create_sheet("Participants")
            ws_students.append([
                "Reg No", "Name", "Course", "Department", "Email", "Phone",
                "Status", "Qualified Round", "Qualified Round Name",
            ])
            round_name_map = {int(r.get("round_order") or 0): r.get("round_name") for r in rounds}
            for s in students:
                q_round = int(s.get("current_round") or 0)
                ws_students.append([
                    s.get("reg_no"),
                    s.get("student_name"),
                    s.get("course"),
                    s.get("department"),
                    s.get("email"),
                    s.get("mobile_number"),
                    s.get("status"),
                    q_round,
                    round_name_map.get(q_round, "Not Cleared" if q_round == 0 else ""),
                ])

            for ws in (ws_info, ws_rounds, ws_students):
                for col in ws.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 42)

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(
                buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"drive_{drive_id}_progress.xlsx",
            )
        except Error as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/api/cdm/drive/<int:drive_id>/export.pdf")
    def cdm_drive_export_pdf(drive_id):
        try:
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas
            except ImportError:
                return jsonify({"error": "PDF export requires reportlab package."}), 500

            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                "SELECT d.*, c.company_name "
                "FROM company_drives d JOIN companies c ON d.company_id = c.company_id "
                "WHERE d.drive_id=%s",
                (drive_id,),
            )
            drive = cursor.fetchone()
            if not drive:
                cursor.close()
                conn.close()
                return jsonify({"error": "Drive not found"}), 404

            cursor.execute(
                "SELECT round_order, round_name FROM drive_rounds WHERE drive_id=%s ORDER BY round_order",
                (drive_id,),
            )
            rounds = cursor.fetchall()

            cursor.execute(
                "SELECT ds.reg_no, ds.status, ds.current_round, s.student_name "
                "FROM drive_students ds JOIN students s ON ds.reg_no=s.reg_no "
                "WHERE ds.drive_id=%s ORDER BY s.student_name",
                (drive_id,),
            )
            students = cursor.fetchall()
            cursor.close()
            conn.close()

            buf = BytesIO()
            pdf = canvas.Canvas(buf, pagesize=A4)
            width, height = A4
            y = height - 36

            def write_line(text, step=14):
                nonlocal y
                if y < 50:
                    pdf.showPage()
                    y = height - 36
                pdf.drawString(36, y, str(text))
                y -= step

            write_line(f"Drive Progress Report - #{drive_id}", 18)
            write_line(f"Company: {drive.get('company_name')}")
            write_line(f"Role: {drive.get('role') or '—'}")
            write_line(f"CTC: {drive.get('ctc_text') or '—'}")
            write_line(f"Status: {drive.get('status') or 'Upcoming'}")
            write_line(f"Process Date: {drive.get('process_date') or '—'}")
            write_line("", 8)

            write_line("Rounds:", 16)
            if not rounds:
                write_line("- No rounds added")
            for r in rounds:
                order = int(r.get("round_order") or 0)
                qualified = len([s for s in students if int(s.get("current_round") or 0) >= order])
                write_line(f"- R{order} {r.get('round_name')}: qualified {qualified}")

            write_line("", 8)
            write_line("Participants:", 16)
            if not students:
                write_line("- No participants linked")
            for s in students:
                write_line(
                    f"- {s.get('student_name')} ({s.get('reg_no')}): "
                    f"Status={s.get('status')}, Qualified Round={int(s.get('current_round') or 0)}"
                )

            pdf.save()
            buf.seek(0)
            return send_file(
                buf,
                mimetype="application/pdf",
                as_attachment=True,
                download_name=f"drive_{drive_id}_progress.pdf",
            )
        except Error as e:
            return jsonify({"error": str(e)}), 500

    # ── Round Management API ─────────────────────────────────────────────
    @app.route("/api/cdm/drive/<int:drive_id>/rounds", methods=["GET"])
    def cdm_drive_rounds(drive_id):
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                "SELECT * FROM drive_rounds WHERE drive_id = %s ORDER BY round_order",
                (drive_id,),
            )
            rows = cursor.fetchall()
            normalize_rows(rows)
            cursor.close()
            conn.close()
            return jsonify({"rounds": rows})
        except Error as e:
            return jsonify({"rounds": [], "error": str(e)}), 500

    @app.route("/api/cdm/drive/<int:drive_id>/rounds", methods=["POST"])
    def cdm_add_round(drive_id):
        data = request.get_json()
        if not data or "round_name" not in data:
            return jsonify({"error": "round_name is required"}), 400
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT IFNULL(MAX(round_order), 0) + 1 AS next_order FROM drive_rounds WHERE drive_id = %s",
                (drive_id,),
            )
            next_order = cursor.fetchone()[0]
            round_date = parse_date_field(data.get("round_date"))
            cursor.execute(
                "INSERT INTO drive_rounds (drive_id, round_name, round_order, round_date) VALUES (%s, %s, %s, %s)",
                (drive_id, data["round_name"], next_order, round_date),
            )
            round_id = cursor.lastrowid
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True, "round_id": round_id, "round_order": next_order})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/cdm/round/<int:round_id>", methods=["DELETE"])
    def cdm_delete_round(round_id):
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            actor = get_request_actor()
            cursor.execute(
                "SELECT drive_id, round_order FROM drive_rounds WHERE round_id=%s",
                (round_id,),
            )
            row = cursor.fetchone()
            if not row:
                cursor.close()
                conn.close()
                return jsonify({"ok": False, "error": "Round not found"}), 404

            drive_id = row["drive_id"]
            deleted_order = int(row["round_order"] or 0)

            cursor.execute(
                "SELECT reg_no, current_round, status FROM drive_students "
                "WHERE drive_id=%s AND current_round >= %s",
                (drive_id, deleted_order),
            )
            before_rows = cursor.fetchall()

            cursor.execute("DELETE FROM drive_rounds WHERE round_id=%s", (round_id,))

            cursor.execute(
                "UPDATE drive_rounds SET round_order = round_order - 1 "
                "WHERE drive_id=%s AND round_order > %s",
                (drive_id, deleted_order),
            )

            cursor.execute(
                "UPDATE drive_students "
                "SET current_round = CASE "
                "  WHEN current_round > %s THEN current_round - 1 "
                "  WHEN current_round = %s THEN GREATEST(%s, 0) "
                "  ELSE current_round END "
                "WHERE drive_id=%s",
                (deleted_order, deleted_order, deleted_order - 1, drive_id),
            )
            cursor.execute(
                "UPDATE drive_students SET current_round = 0 "
                "WHERE drive_id=%s AND (current_round IS NULL OR current_round < 0)",
                (drive_id,),
            )

            if before_rows:
                cursor.execute(
                    "SELECT reg_no, current_round, status FROM drive_students "
                    "WHERE drive_id=%s",
                    (drive_id,),
                )
                after_map = {r["reg_no"]: r for r in cursor.fetchall()}
                for before in before_rows:
                    reg_no = before["reg_no"]
                    after = after_map.get(reg_no)
                    if not after:
                        continue
                    from_round = int(before.get("current_round") or 0)
                    to_round = int(after.get("current_round") or 0)
                    if from_round != to_round:
                        append_round_transition_log(
                            cursor,
                            drive_id,
                            reg_no,
                            from_round,
                            to_round,
                            before.get("status"),
                            after.get("status"),
                            actor,
                            note="round_deleted",
                        )

            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    # ── Calendar Data API ────────────────────────────────────────────────
    @app.route("/api/cdm/calendar")
    def cdm_calendar():
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT d.drive_id, d.company_id, c.company_name, d.role, d.ctc_text,
                      d.process_date, d.jd_received_date,
                       d.status, d.location
                FROM company_drives d
                JOIN companies c ON d.company_id = c.company_id
                WHERE d.process_date IS NOT NULL
                ORDER BY d.process_date
            """)
            rows = cursor.fetchall()
            normalize_rows(rows)
            for row in rows:
                process_date = row.get("process_date")
                row["process_date_key"] = str(process_date)[:10] if process_date else None
            cursor.close()
            conn.close()
            return jsonify({"events": rows})
        except Error as e:
            return jsonify({"events": [], "error": str(e)}), 500

    @app.route("/api/cdm/alerts")
    def cdm_alerts():
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                "SELECT d.drive_id, d.company_id, c.company_name, d.role, d.process_date, d.status, "
                "COUNT(DISTINCT h.hr_id) AS hr_count, "
                "COUNT(DISTINCT r.round_id) AS rounds_count, "
                "MAX(r.round_date) AS last_round_date "
                "FROM company_drives d "
                "JOIN companies c ON d.company_id = c.company_id "
                "LEFT JOIN company_hr h ON h.company_id = c.company_id "
                "LEFT JOIN drive_rounds r ON r.drive_id = d.drive_id "
                "GROUP BY d.drive_id, d.company_id, c.company_name, d.role, d.process_date, d.status "
                "ORDER BY d.drive_id DESC"
            )
            rows = cursor.fetchall()
            normalize_rows(rows)
            cursor.close()
            conn.close()

            today = date.today()
            alerts = []
            for row in rows:
                status = (row.get("status") or "Upcoming").strip()
                status_l = status.lower()
                process_date_str = row.get("process_date")
                last_round_date_str = row.get("last_round_date")
                hr_count = int(row.get("hr_count") or 0)
                rounds_count = int(row.get("rounds_count") or 0)

                flags = []

                if not process_date_str:
                    flags.append("missing_process_date")

                if hr_count == 0:
                    flags.append("missing_hr")

                process_date_obj = None
                last_round_obj = None
                if process_date_str:
                    try:
                        process_date_obj = datetime.strptime(str(process_date_str), "%Y-%m-%d").date()
                    except ValueError:
                        process_date_obj = None
                if last_round_date_str:
                    try:
                        last_round_obj = datetime.strptime(str(last_round_date_str), "%Y-%m-%d").date()
                    except ValueError:
                        last_round_obj = None

                if status_l not in ("completed", "cancelled"):
                    if rounds_count > 0 and last_round_obj and last_round_obj < today:
                        flags.append("rounds_overdue")
                    if process_date_obj and process_date_obj < today and rounds_count == 0:
                        flags.append("stale_no_rounds")

                if flags:
                    alerts.append({
                        "drive_id": row.get("drive_id"),
                        "company_id": row.get("company_id"),
                        "company_name": row.get("company_name"),
                        "role": row.get("role"),
                        "process_date": process_date_str,
                        "status": status,
                        "flags": flags,
                    })

            return jsonify({
                "alerts": alerts,
                "counts": {
                    "total": len(alerts),
                    "missing_process_date": len([a for a in alerts if "missing_process_date" in a["flags"]]),
                    "missing_hr": len([a for a in alerts if "missing_hr" in a["flags"]]),
                    "rounds_overdue": len([a for a in alerts if "rounds_overdue" in a["flags"]]),
                    "stale_no_rounds": len([a for a in alerts if "stale_no_rounds" in a["flags"]]),
                },
            })
        except Error as e:
            return jsonify({"alerts": [], "error": str(e)}), 500

    # ── Placement Source Analytics ────────────────────────────────────────
    @app.route("/api/cdm/placement-sources")
    def cdm_placement_sources():
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT s.company_name, COUNT(*) AS placed_count,
                       ROUND(AVG(s.ctc), 2) AS avg_ctc,
                       MAX(s.ctc) AS max_ctc,
                       GROUP_CONCAT(DISTINCT s.course SEPARATOR ', ') AS courses
                FROM students s
                WHERE s.status = 'Placed' AND s.company_name IS NOT NULL AND s.company_name != ''
                GROUP BY s.company_name
                ORDER BY placed_count DESC
            """)
            rows = cursor.fetchall()
            normalize_rows(rows)

            cursor.execute("""
                SELECT c.company_name, COUNT(*) AS linked_count
                FROM drive_students ds
                JOIN company_drives d ON ds.drive_id = d.drive_id
                JOIN companies c ON d.company_id = c.company_id
                WHERE ds.status IN ('Selected', 'Placed')
                GROUP BY c.company_name
            """)
            linked = {r["company_name"]: r["linked_count"] for r in cursor.fetchall()}

            for r in rows:
                r["linked_placed"] = linked.get(r["company_name"], 0)

            cursor.close()
            conn.close()
            return jsonify({"sources": rows})
        except Error as e:
            return jsonify({"sources": [], "error": str(e)}), 500

    # ── Unique courses & departments from students table ────────────────
    @app.route("/api/cdm/unique-courses-departments")
    def cdm_unique_courses_departments():
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT DISTINCT course FROM students WHERE course IS NOT NULL AND course != '' ORDER BY course"
            )
            courses = [r[0] for r in cursor.fetchall()]
            cursor.execute(
                "SELECT DISTINCT department FROM students WHERE department IS NOT NULL AND department != '' ORDER BY department"
            )
            departments = [r[0] for r in cursor.fetchall()]
            cursor.close()
            conn.close()
            return jsonify({"courses": courses, "departments": departments})
        except Error as e:
            return jsonify({"courses": [], "departments": [], "error": str(e)}), 500

    # ── Course presets CRUD ─────────────────────────────────────────────
    @app.route("/api/cdm/course-presets")
    def cdm_list_presets():
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM course_presets ORDER BY preset_name")
            presets = cursor.fetchall()
            preset_ids = [p["preset_id"] for p in presets]
            items_map = {}
            if preset_ids:
                fmt = ",".join(["%s"] * len(preset_ids))
                cursor.execute(
                    f"SELECT preset_id, course_name FROM course_preset_items WHERE preset_id IN ({fmt})",
                    tuple(preset_ids),
                )
                for row in cursor.fetchall():
                    items_map.setdefault(row["preset_id"], []).append(row["course_name"])
            for p in presets:
                p["courses"] = items_map.get(p["preset_id"], [])
            cursor.close()
            conn.close()
            return jsonify({"presets": presets})
        except Error as e:
            return jsonify({"presets": [], "error": str(e)}), 500

    @app.route("/api/cdm/course-presets", methods=["POST"])
    def cdm_create_preset():
        data = request.get_json(silent=True) or {}
        name = (data.get("preset_name") or "").strip()
        courses = data.get("courses") or []
        if not name:
            return jsonify({"ok": False, "error": "Preset name is required."}), 400
        if not isinstance(courses, list) or not courses:
            return jsonify({"ok": False, "error": "At least one course is required."}), 400
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO course_presets (preset_name) VALUES (%s)", (name,)
            )
            pid = cursor.lastrowid
            for c in courses:
                c = str(c).strip()
                if c:
                    cursor.execute(
                        "INSERT IGNORE INTO course_preset_items (preset_id, course_name) VALUES (%s, %s)",
                        (pid, c),
                    )
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True, "preset_id": pid, "preset_name": name, "courses": courses})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/cdm/course-presets/<int:preset_id>", methods=["DELETE"])
    def cdm_delete_preset(preset_id):
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM course_presets WHERE preset_id=%s", (preset_id,))
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"ok": True})
        except Error as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    # ── Company courses/departments fetch ────────────────────────────────
    @app.route("/api/cdm/company/<company_id>/details")
    def cdm_company_details_api(company_id):
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM companies WHERE company_id=%s", (company_id,))
            company = cursor.fetchone()
            if not company:
                cursor.close()
                conn.close()
                return jsonify({"error": "Company not found."}), 404
            company["courses"] = []
            company["departments"] = []
            cursor.close()
            conn.close()
            return jsonify(company)
        except Error as e:
            return jsonify({"error": str(e)}), 500

    # ── Student Search (for linking UI) ──────────────────────────────────
    @app.route("/api/cdm/search-students")
    def cdm_search_students():
        q = request.args.get("q", "").strip()
        if len(q) < 2:
            return jsonify({"results": []})
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            like = f"%{q}%"
            cursor.execute(
                "SELECT reg_no, student_name, course, department, status "
                "FROM students WHERE "
                "student_name LIKE %s OR reg_no LIKE %s OR course LIKE %s "
                "ORDER BY student_name LIMIT 15",
                (like, like, like),
            )
            rows = cursor.fetchall()
            cursor.close()
            conn.close()
            return jsonify({"results": rows})
        except Error as e:
            return jsonify({"results": [], "error": str(e)}), 500
